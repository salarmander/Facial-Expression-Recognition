import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from rich import print

weight_border = Border(
    left  = Side(style="thin"),
    right = Side(style="thin"),
    top   = Side(style="thin"),
    bottom= Side(style="thin")
)

class ExportToExcel:
    def __init__(self, excel_path="fer_report.xlsx"):
        self.excel_path = excel_path

        if not os.path.exists(self.excel_path):
            wb = Workbook()
            ws = wb.active
            main_offset_col = 2
            sub_offset_col  = 8

            main_headers = ["No", "Inputs", "Emotion", "Result", "Percent"]
            for idx, header in enumerate(main_headers):
                cell           = ws.cell(row=2, column=idx + main_offset_col, value=header)
                cell.font      = Font(bold=True, size=13, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill      = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.border    = weight_border

            sub_headers = ["Overall", "Angry", "Disgust", "Fear", "Happy", "Neutral", "Sad", "Surprise"]
            for idx, header in enumerate(sub_headers):
                cell           = ws.cell(row=2 , column=idx + sub_offset_col, value=header)
                cell.font      = Font(bold=True, size=13, color="000000")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = weight_border           

            wb.save(self.excel_path)

    def export(self, result, img):
        image_path = img['img_path']

        wb  = load_workbook(self.excel_path)
        ws  = wb.active
        row = ws.max_row + 1
        ws.cell(row=row, column=2, value=row - 2)
        ws.cell(row=row, column=3, value=image_path)        
        ws.cell(row=row, column=4, value=img['emotion'])
        ws.cell(row=row, column=5, value=result['dominant'])
        ws.cell(row=row, column=6, value=result['percent'])

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        emotion_colors = {
            "angry"   : "C6EFCE",
            "disgust" : "FFC7CE",
            "fear"    : "FFEB9C",
            "happy"   : "D9D9D9",
            "neutral" : "F9DFA5",
            "sad"     : "B267E6",
            "surprise": "7EC699",
        }

        emotion    = str(img['emotion']).lower()
        fill_color = emotion_colors.get(emotion, "FFFFFF")
        for col in range(1, 6):
            ws.cell(row, column=col + 1).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            ws.cell(row, column=col + 1).border = weight_border

        # if os.path.exists(image_path):
        #     col_letter = get_column_letter(3)
        #     ws[f"{col_letter}{row}"].hyperlink = image_path
        #     ws[f"{col_letter}{row}"].style = "Hyperlink"
        # else:
        #     ws.cell(row=row, column=3, value=image_path)

        wb.save(self.excel_path)
        print(f"[bright_green][INFO] Added success {img['file_name']} → {row} [/]")

    def export_accuracy(self, accuracy):
        
        wb  = load_workbook(self.excel_path)
        ws  = wb.active
        for idx, emo in enumerate(accuracy):
            ws.cell(row=3, column=idx + 8, value=f"{accuracy[emo]}%")
            ws.cell(row=3, column=idx + 8).alignment = Alignment(horizontal="right", vertical="center")

        emotion_colors = {
            "overall" : "666666",
            "angry"   : "C6EFCE",
            "disgust" : "FFC7CE",
            "fear"    : "FFEB9C",
            "happy"   : "D9D9D9",
            "neutral" : "F9DFA5",
            "sad"     : "B267E6",
            "surprise": "7EC699",
        }

        for idx, emo in enumerate(accuracy):
            emo_lower  = str(emo).lower()
            fill_color = emotion_colors.get(emo_lower, "FFFFFF")
            col_index  = idx + 8
            col_letter = get_column_letter(col_index)

            ws.cell(row=2,  column=col_index).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            ws.cell(row=3, column=col_index).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            ws.cell(row=3, column=col_index).border = weight_border

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
                        
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2
        
        wb.save(self.excel_path)
        print(f"[bright_green][INFO] Successful exports accuracy[/]")